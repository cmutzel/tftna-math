# -*- coding: utf-8 -*-
"""
weekly-activity-totals.py

Created on Tue Feb  7 19:08:14 2017

@author: chrismutzel
"""
import os
import pandas as pd
from openpyxl import load_workbook
from src.products.data_product import DataProduct

import logging
logger = logging.getLogger(__name__)
logger.setLevel(logging.DEBUG)

"""make this visible at the top of the file for easier reading"""
PRODUCT_NAME = "base_training_goals"

def read_table(sheet, col_names, start_row, end_row):
    """
        There are two tables in the training goals spreadsheet
        They can be read the same way, this just makes it easy
        
        Params
        -------
        int start_row: integar representing the first row that should be read
        int end_row: integer representing the last row that should be read
        
        Returns
        -------
        Pandas.DataFrame with week # as index (i.e. 1-N).  Frame includes a column
        with sum of hours for each activity each week and also a column showing
        total of training hours. 
        i.e. 
        week_number  week_starting   zone_1,2   zone_3    zone4      ... total
                  1      2/29/2016        0.5        1        0      ...   1.5 
                  2      3/xx/2016          0        1        1      ...     2 
        ...
        
    """ 
    names = []
    data = []
    for name, loc in col_names.iteritems(): 
        names.append(name)
        col_values = []
        col_ind = ord(loc) - 64
        for cells in sheet.iter_cols(
                min_col=col_ind, max_col=col_ind,
                min_row=start_row, max_row=end_row):
            for cell in cells:
                col_values.append(cell.value)
        data.append(col_values)
        
    final = dict(zip(names, data))
    d = pd.DataFrame(data=final)
    d["week_number"] = d["week_number"].astype("int")
    
    return d.set_index("week_number")
    

class BaseTrainingGoals(DataProduct):
    product_name = PRODUCT_NAME
    requires = []

    def __init__(self, force_build_all=False):
        super(BaseTrainingGoals, self).__init__(
                BaseTrainingGoals.product_name)
        
    def build(self):
        """
        Reads training goals (total hours, and percentages of each type of 
            training activity from previously compiled spreadsheet) 
    
        Parameters
        ----------
        None
    
        Returns
        -------
        Pandas.DataFrame with week # as index (i.e. 1-N).  Frame includes a column
        with sum of hours for each activity each week and also a column showing
        total of training hours. 
        i.e. 
        week_number  week_starting   zone_1,2   zone_3    zone4      ... total
                  1      2/29/2016        0.5        1        0      ...   1.5 
                  2      3/xx/2016          0        1        1      ...     2 
        ...
        
        """
        path_to_sheet = os.path.join(os.path.dirname(
                os.path.abspath(__file__)),
                "../../data/raw/Chris - Training Goals 2016.xlsx"
        )
        
        wb = load_workbook(path_to_sheet)
        sheet = wb["Chris - Base period"]
            
        col_wks_1_8 = {
            "week_number":  "A",
            "Date":         "B",
            "Period":       "D",
            "total":        "E",
            "zone_1_per":   "F",
            "zone_2_per":   "H",
            "zone_3_per":   "J",
            "zone_max_strength_per": "L",
            "zone_cragging_per":     "P",
            "zone_alpine_climbing_per": "N"
        }

        col_wks_9_16 = {
            "week_number":  "A",
            "Date":         "B",
            "Period":       "D",
            "total":        "E",
            "zone_1_per":   "F",
            "zone_2_per":   "H",
            "zone_3_per":   "J",
            "zone_strength_per": "L",
            "zone_cragging_per":     "P",
            "zone_alpine_climbing_per": "N"
        }
        
        weekly_training_goals = pd.concat([
                read_table(sheet, col_wks_1_8, 5, 12),
                read_table(sheet, col_wks_9_16, 23, 30),
                            ])
        all_keys = set(col_wks_1_8.keys()).union(set(col_wks_9_16.keys()))
        print all_keys
        for name in all_keys:
            print name
            if "zone_" in name:                
                weekly_training_goals[name[0:-4]] = (
                    weekly_training_goals[name]*weekly_training_goals["total"])
                del weekly_training_goals[name]
        
        return weekly_training_goals.sort_index().fillna(value=0)
    

    def verify(self):
        """ We don't do any checks on this data *yet* """
        
        return self.value
    
if __name__ == "__main__":
    from src.products.product_factory import DataProductFactory
    factory = DataProductFactory()
    value = factory.get_product(PRODUCT_NAME, force_build=True)
    print value